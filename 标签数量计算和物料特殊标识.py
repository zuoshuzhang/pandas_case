import openpyxl
from openpyxl.styles import Font, Color, PatternFill

# 读取工作簿
file_path = r'NW-SH-20240529.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# 定义字体和填充样式
red_font = Font(color=Color(rgb='FFFF0000'))
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_font = Font(color='008000')

# 锁定起始单元格
current_cell = ws['B3'].value
current_count = ws['B2'].value
current_sum = 0
for i in range(3, ws.max_row - 6, 10):
    # 获取物料号码
    cell_value = ws.cell(row=i, column=2).value
    if cell_value != current_cell:
        ws.cell(row=i, column=2).font = red_font
        ws.cell(row=i, column=2).fill = yellow_fill
        ws.cell(row=i + 6, column=2).value = ws.cell(row=i-1, column=2).value
        current_sum = ws.cell(row=i - 1, column=2).value
    else:
        current_count = ws.cell(row=i - 1, column=2).value
        ws.cell(row=i + 6, column=2).value = current_sum + current_count
        current_sum = ws.cell(row=i + 6, column=2).value
    current_cell = cell_value

# 保存文件
wb.save(file_path)
